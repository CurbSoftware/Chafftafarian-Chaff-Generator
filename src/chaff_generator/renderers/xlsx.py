"""XLSX renderer — openpyxl over TabularDocument (spec sections 66, 74).

Typed cells (dates as dates, currency as numbers with a currency format),
styled headers, an occasional ``=SUM`` total row, and ``write_only`` mode
for large targets so the workbook never fully materializes in memory.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from chaff_generator.core.hashing import hash_file
from chaff_generator.renderers.base import RendererCapabilities, RenderResult

if TYPE_CHECKING:
    from pathlib import Path
    from typing import Any

    from chaff_generator.content.context import RenderContext
    from chaff_generator.renderers.base import Renderer
    from chaff_generator.renderers.documents import SemanticDocument, Sheet, TabularDocument

CAPABILITIES = RendererCapabilities(
    extension="xlsx",
    supports_exact_size=False,
    supports_target_size=True,
    supports_streaming=False,
    semantic_document="tabular",
    size_category="spreadsheet",
)

#: Above this target, build the workbook in write-only mode (streamed XML).
WRITE_ONLY_THRESHOLD = 8 << 20

_CURRENCY_FORMAT = '"$"#,##0.00'
_DATE_FORMAT = "YYYY-MM-DD"


class XlsxRenderer:
    id = "xlsx"
    capabilities = CAPABILITIES

    def render(
        self,
        document: SemanticDocument | None,
        destination: Path,
        context: RenderContext,
    ) -> RenderResult:
        from openpyxl import Workbook

        if document is None:
            document = self._fallback_document(context)
        from chaff_generator.renderers.documents import TabularDocument

        assert isinstance(document, TabularDocument)
        if not document.sheets:
            self._fail("spreadsheet document has no sheets")

        write_only = context.desired_size > WRITE_ONLY_THRESHOLD
        workbook = Workbook(write_only=write_only)

        for position, sheet in enumerate(document.sheets):
            worksheet = self._new_sheet(workbook, sheet, position, write_only)
            self._write_sheet(worksheet, sheet, write_only)

        workbook.save(str(destination))
        size = destination.stat().st_size
        return RenderResult(
            path=destination,
            size=size,
            renderer_id=self.id,
            template_id=context.template_id,
            sha256=hash_file(destination),
        )

    # ---------------------------------------------------------------- internals

    def _fail(self, reason: str) -> None:
        from chaff_generator.core.errors import RendererError

        raise RendererError(reason)

    def _fallback_document(self, context: RenderContext) -> TabularDocument:
        """No template: materialize a bounded TabularDocument from synthetic rows."""
        from chaff_generator.renderers.documents import Sheet, TabularDocument
        from chaff_generator.renderers.tabular import _synthetic_rows

        title, columns, rows = _synthetic_rows(context)
        limit = 20_000
        materialized = [list(row) for _, row in zip(range(limit), rows, strict=False)]
        return TabularDocument(
            title=title,
            author=context.world.primary_user.full_name,
            sheets=[Sheet(name=(title or "Data")[:31], columns=list(columns), rows=materialized)],
        )

    def _new_sheet(self, workbook: Any, sheet: Sheet, position: int, write_only: bool) -> Any:
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_align = Alignment(horizontal="center")

        if write_only:
            worksheet = workbook.create_sheet(title=sheet.name[:31])
            cells = []
            for name in sheet.columns:
                cell = WriteOnlyCell(worksheet, value=name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cells.append(cell)
            worksheet.append(cells)
            return worksheet

        worksheet = workbook.active if position == 0 else workbook.create_sheet()
        worksheet.title = sheet.name[:31]
        worksheet.append(list(sheet.columns))
        for column in range(1, len(sheet.columns) + 1):
            cell = worksheet.cell(row=1, column=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            from openpyxl.utils import get_column_letter

            worksheet.column_dimensions[get_column_letter(column)].width = max(
                12, min(32, len(str(sheet.columns[column - 1])) + 4)
            )
        return worksheet

    def _write_sheet(self, worksheet: Any, sheet: Sheet, write_only: bool) -> None:
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        for row in sheet.rows:
            typed = self._typed_row(sheet, row)
            if write_only:
                cells = []
                for index, value in enumerate(typed):
                    if index in sheet.currency_columns and isinstance(value, (int, float)):
                        cell = WriteOnlyCell(worksheet, value=value)
                        cell.number_format = _CURRENCY_FORMAT
                        cells.append(cell)
                    else:
                        cells.append(value)
                worksheet.append(cells)
            else:
                worksheet.append(typed)

        if write_only:
            return

        from datetime import date as date_type

        for column in sheet.currency_columns:
            letter = get_column_letter(column + 1)
            for row_index in range(2, len(sheet.rows) + 2):
                cell = worksheet[f"{letter}{row_index}"]
                cell.number_format = _CURRENCY_FORMAT
        for column in sheet.date_columns:
            letter = get_column_letter(column + 1)
            for row_index in range(2, len(sheet.rows) + 2):
                cell = worksheet[f"{letter}{row_index}"]
                if isinstance(cell.value, date_type):
                    cell.number_format = _DATE_FORMAT

        if sheet.total_row and sheet.rows:
            last = len(sheet.rows) + 2
            totals = [""] * len(sheet.columns)
            totals[0] = "Total"
            for column in sorted(sheet.currency_columns):
                letter = get_column_letter(column + 1)
                totals[column] = f"=SUM({letter}2:{letter}{last - 1})"
            worksheet.append(totals)
            for column in range(len(sheet.columns)):
                worksheet.cell(row=last, column=column + 1).font = Font(bold=True)

    def _typed_row(self, sheet: Sheet, row: list[Any]) -> list[Any]:
        """Coerce row values to native types (dates, currency floats)."""
        from datetime import date

        typed: list[Any] = []
        for index, value in enumerate(row):
            converted = value
            if isinstance(value, str):
                text = value.strip()
                if index in sheet.date_columns:
                    try:
                        converted = date.fromisoformat(text)
                    except ValueError:
                        converted = value
                elif index in sheet.currency_columns:
                    cleaned = text.replace("$", "").replace(",", "")
                    try:
                        converted = float(cleaned)
                    except ValueError:
                        converted = value
            typed.append(converted)
        return typed


def get_renderer(renderer_id: str) -> Renderer:
    if renderer_id != "xlsx":
        raise ValueError(f"xlsx module cannot serve renderer id {renderer_id!r}")
    return XlsxRenderer()
